"""
excel_extraction.py
---------------------
Extract structured data from Excel/CSV files using:
    - pandas   : reading sheets/CSVs into DataFrames, dtype inference, JSON export
    - openpyxl : low-level access (cell comments, merged cells, formulas, sheet metadata)

Input : ../data/excel/*.xlsx, *.xls, *.csv
Output:
    ../output/json/<name>_excel.json
    ../output/dataframes/<name>_<sheet>.csv        (one CSV per sheet, cleaned)
    ../output/key_value/<name>_excel.txt           (workbook/sheet metadata as key-value)

Install:
    pip install pandas openpyxl
"""

import json
from pathlib import Path

import pandas as pd
import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
INPUT_DIR = BASE_DIR / "data" / "excel"
OUTPUT_JSON_DIR = BASE_DIR / "output" / "json"
OUTPUT_DF_DIR = BASE_DIR / "output" / "dataframes"
OUTPUT_KV_DIR = BASE_DIR / "output" / "key_value"

for d in (OUTPUT_JSON_DIR, OUTPUT_DF_DIR, OUTPUT_KV_DIR):
    d.mkdir(parents=True, exist_ok=True)


def extract_sheet_metadata_openpyxl(filepath: Path):
    """Use openpyxl for workbook-level metadata: sheet names, dims, merged cells."""
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    meta = {
        "sheet_names": wb.sheetnames,
        "sheets": {}
    }
    for name in wb.sheetnames:
        ws = wb[name]
        meta["sheets"][name] = {
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
        }
    return meta


def process_excel_file(filepath: Path):
    all_sheets = pd.read_excel(str(filepath), sheet_name=None)  # dict of {sheet_name: DataFrame}
    meta = extract_sheet_metadata_openpyxl(filepath)

    structured = {
        "source_file": filepath.name,
        "workbook_metadata": meta,
        "sheets": {},
    }

    kv_pairs = {
        "num_sheets": len(all_sheets),
        "sheet_names": ", ".join(all_sheets.keys()),
    }

    for sheet_name, df in all_sheets.items():
        df = df.dropna(how="all")  # drop fully-empty rows
        structured["sheets"][sheet_name] = {
            "num_rows": len(df),
            "num_columns": len(df.columns),
            "columns": list(df.columns.astype(str)),
            "records": json.loads(df.to_json(orient="records", date_format="iso")),
        }

        # ---- Save each sheet as its own CSV ----
        safe_sheet_name = sheet_name.replace("/", "_").replace(" ", "_")
        df_path = OUTPUT_DF_DIR / f"{filepath.stem}_{safe_sheet_name}.csv"
        df.to_csv(df_path, index=False)

        kv_pairs[f"sheet::{sheet_name}::rows"] = len(df)
        kv_pairs[f"sheet::{sheet_name}::columns"] = len(df.columns)

    # ---- Save JSON ----
    json_path = OUTPUT_JSON_DIR / f"{filepath.stem}_excel.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(structured, f, indent=2, ensure_ascii=False)

    # ---- Save key-value text ----
    kv_path = OUTPUT_KV_DIR / f"{filepath.stem}_excel.txt"
    with open(kv_path, "w", encoding="utf-8") as f:
        for k, v in kv_pairs.items():
            f.write(f"{k}: {v}\n")

    print(f"[OK] Processed {filepath.name} -> {json_path.name}")
    return structured


def process_csv_file(filepath: Path):
    df = pd.read_csv(str(filepath))
    df = df.dropna(how="all")

    structured = {
        "source_file": filepath.name,
        "num_rows": len(df),
        "num_columns": len(df.columns),
        "columns": list(df.columns.astype(str)),
        "records": json.loads(df.to_json(orient="records", date_format="iso")),
    }

    json_path = OUTPUT_JSON_DIR / f"{filepath.stem}_csv.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(structured, f, indent=2, ensure_ascii=False)

    df_path = OUTPUT_DF_DIR / f"{filepath.stem}_csv.csv"
    df.to_csv(df_path, index=False)

    kv_path = OUTPUT_KV_DIR / f"{filepath.stem}_csv.txt"
    with open(kv_path, "w", encoding="utf-8") as f:
        f.write(f"num_rows: {len(df)}\n")
        f.write(f"num_columns: {len(df.columns)}\n")
        f.write(f"columns: {', '.join(df.columns.astype(str))}\n")

    print(f"[OK] Processed {filepath.name} -> {json_path.name}")
    return structured


def main():
    excel_files = list(INPUT_DIR.glob("*.xlsx")) + list(INPUT_DIR.glob("*.xls"))
    csv_files = list(INPUT_DIR.glob("*.csv"))

    if not excel_files and not csv_files:
        print(f"No Excel/CSV files found in {INPUT_DIR}")
        return

    for filepath in excel_files:
        try:
            process_excel_file(filepath)
        except Exception as e:
            print(f"[ERROR] Failed to process {filepath.name}: {e}")

    for filepath in csv_files:
        try:
            process_csv_file(filepath)
        except Exception as e:
            print(f"[ERROR] Failed to process {filepath.name}: {e}")


if __name__ == "__main__":
    main()